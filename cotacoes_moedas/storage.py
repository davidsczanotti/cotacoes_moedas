from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import csv
import re
from typing import Callable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .bcb_ptax import PtaxQuote
from .investing import Quote
from .valor_globo import BidAskQuote


_DECIMAL_4 = Decimal("0.0001")
_DECIMAL_10 = Decimal("0.0000000001")
_DATE_NUMBER_FORMAT = "dd/mm/yyyy"
_QUOTE_NUMBER_FORMAT = "0.0000"
_PERCENT_NUMBER_FORMAT = "0.00%"
_CDI_NUMBER_FORMAT = "0.0000000000"
_DATE_PATTERN = re.compile(r"^\d{2}/\d{2}/\d{4}$")
_LOCAL_TZ = datetime.now().astimezone().tzinfo or timezone.utc
_LOG_COLUMN_INDEX = 15
_TJLP_COLUMN = "L"
_SELIC_COLUMN = "M"
_CDI_COLUMN = "N"
_LOG_COLUMN = "O"
_FIRST_COLUMN = "A"
_LAST_COLUMN_INDEX = 15
_DEFAULT_CSV_HEADER = [
    "Data",
    "Dolar Oficial Compra",
    "Dolar Oficial Venda",
    "Dolar PTAX Compra",
    "Dolar PTAX Venda",
    "Dolar Turismo Compra",
    "Dolar Turismo Venda",
    "Euro PTAX Compra",
    "Euro PTAX Venda",
    "CHF PTAX Compra",
    "CHF PTAX Venda",
    "TJLP",
    "SELIC",
    "CDI",
    "Situacao",
]
_COLUMN_WIDTHS = {
    "A": 10.5,
    "B": 12.5,
    "C": 11.5,
    "D": 12.5,
    "E": 11.5,
    "F": 12.5,
    "G": 11.5,
    "H": 11.5,
    "I": 11.5,
    "J": 11.5,
    "K": 11.5,
    "L": 10.0,
    "M": 10.0,
    "N": 12.5,
    "O": 30.0,
}
_HEADER_TOP_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
_ROW_ODD_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF")
_ROW_EVEN_FILL = PatternFill(fill_type="solid", fgColor="ECF3FB")
_BORDER_SIDE = Side(style="thin", color="9CB6D9")
_CELL_BORDER = Border(
    left=_BORDER_SIDE,
    right=_BORDER_SIDE,
    top=_BORDER_SIDE,
    bottom=_BORDER_SIDE,
)
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="1F2937")
_BODY_FONT = Font(name="Calibri", size=11, bold=False, color="000000")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def _quantize_4(value: Decimal) -> Decimal:
    return value.quantize(_DECIMAL_4, rounding=ROUND_HALF_UP)


def _quantize_10(value: Decimal) -> Decimal:
    return value.quantize(_DECIMAL_10, rounding=ROUND_HALF_UP)


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _set_cell(
    sheet,
    address: str,
    value: object,
    *,
    number_format: str | None = None,
    overwrite: bool = True,
) -> bool:
    cell = sheet[address]
    if not overwrite and not _is_blank(cell.value):
        return False
    cell.value = value
    if number_format and cell.number_format != number_format:
        cell.number_format = number_format
    return True


def _looks_like_log(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = " ".join(value.split()).upper()
    return text.startswith("OK ") or text.startswith("ERRO ")


def _ensure_layout(sheet) -> None:
    # Mantem compatibilidade com planilhas antigas (log na coluna L)
    # e garante os cabecalhos finais na linha 2.
    sheet["L1"] = None
    sheet["M1"] = None
    sheet["N1"] = None
    sheet["O1"] = None
    sheet["L2"] = "TJLP"
    sheet["M2"] = "SELIC"
    sheet["N2"] = "CDI"
    sheet["O2"] = "Situação"

    for row in range(3, sheet.max_row + 1):
        old_log = sheet.cell(row=row, column=12).value
        if not _looks_like_log(old_log):
            continue
        new_log = sheet.cell(row=row, column=_LOG_COLUMN_INDEX).value
        if _is_blank(new_log):
            sheet.cell(row=row, column=_LOG_COLUMN_INDEX).value = str(old_log).strip()
        sheet.cell(row=row, column=12).value = None


def _apply_visual_style(sheet) -> None:
    last_data_row = _find_last_date_row(sheet) or 2
    last_row = max(2, last_data_row)

    # Garantir os agrupamentos visuais da linha 1.
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row == 1 and merged.max_row == 1:
            sheet.unmerge_cells(str(merged))
    for merge_ref in ("B1:C1", "D1:E1", "F1:G1", "H1:I1", "J1:K1"):
        sheet.merge_cells(merge_ref)

    for col_name, width in _COLUMN_WIDTHS.items():
        sheet.column_dimensions[col_name].width = width

    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[2].height = 20
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"{_FIRST_COLUMN}2:{_LOG_COLUMN}{last_row}"

    for col_index in range(1, _LAST_COLUMN_INDEX + 1):
        top_cell = sheet.cell(row=1, column=col_index)
        top_cell.fill = _HEADER_TOP_FILL
        top_cell.font = _HEADER_FONT
        top_cell.alignment = _ALIGN_CENTER
        top_cell.border = _CELL_BORDER

        header_cell = sheet.cell(row=2, column=col_index)
        header_cell.fill = _HEADER_FILL
        header_cell.font = _HEADER_FONT
        header_cell.alignment = _ALIGN_CENTER
        header_cell.border = _CELL_BORDER

    for row in range(3, last_row + 1):
        row_fill = _ROW_EVEN_FILL if row % 2 == 0 else _ROW_ODD_FILL
        for col_index in range(1, _LAST_COLUMN_INDEX + 1):
            cell = sheet.cell(row=row, column=col_index)
            cell.fill = row_fill
            cell.font = _BODY_FONT
            cell.border = _CELL_BORDER
            if col_index == 1:
                cell.alignment = _ALIGN_CENTER
            elif col_index == _LOG_COLUMN_INDEX:
                cell.alignment = _ALIGN_LEFT
            else:
                cell.alignment = _ALIGN_RIGHT


def _find_previous_non_blank_value(
    sheet,
    start_row: int,
    column_index: int,
) -> object | None:
    for row in range(start_row - 1, 2, -1):
        if not _coerce_date(sheet.cell(row=row, column=1).value):
            continue
        value = sheet.cell(row=row, column=column_index).value
        if not _is_blank(value):
            return value
    return None


def _repeat_previous_value_if_blank(
    sheet,
    row: int,
    column_index: int,
    *,
    number_format: str | None = None,
) -> bool:
    cell = sheet.cell(row=row, column=column_index)
    if not _is_blank(cell.value):
        return False

    previous = _find_previous_non_blank_value(sheet, row, column_index)
    if _is_blank(previous):
        return False

    cell.value = previous
    if number_format:
        cell.number_format = number_format
    return True


def _load_and_save_workbook(
    path: Path,
    apply_updates: Callable[[object], object],
) -> object:
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        _ensure_layout(sheet)
        result = apply_updates(sheet)
        _normalize_interest_number_formats(sheet)
        _apply_visual_style(sheet)
        workbook.save(path)
        return result
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def _coerce_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _as_local_date(value: datetime) -> date:
    if value.tzinfo is None:
        localized = value.replace(tzinfo=_LOCAL_TZ)
    else:
        localized = value.astimezone(_LOCAL_TZ)
    return localized.date()


def _as_local_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=_LOCAL_TZ)
    return value.astimezone(_LOCAL_TZ)


def _find_row_by_date(sheet, target_date: date) -> int | None:
    for row in range(3, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        cell_date = _coerce_date(cell_value)
        if cell_date == target_date:
            return row
    return None


def _find_last_date_row(sheet) -> int | None:
    last_date_row = None
    for row in range(3, sheet.max_row + 1):
        if _coerce_date(sheet.cell(row=row, column=1).value):
            last_date_row = row
    return last_date_row


def _find_or_create_row_by_date(sheet, target_date: date) -> int:
    row = _find_row_by_date(sheet, target_date)
    if row is not None:
        return row
    last_date_row = _find_last_date_row(sheet)
    row = (last_date_row or 2) + 1
    sheet[f"A{row}"] = target_date
    sheet[f"A{row}"].number_format = _DATE_NUMBER_FORMAT
    return row


def _normalize_interest_number_formats(sheet) -> None:
    last_row = _find_last_date_row(sheet)
    if last_row is None:
        return

    for row in range(3, last_row + 1):
        if not _coerce_date(sheet.cell(row=row, column=1).value):
            continue

        date_cell = sheet.cell(row=row, column=1)
        if date_cell.number_format != _DATE_NUMBER_FORMAT:
            date_cell.number_format = _DATE_NUMBER_FORMAT

        tjlp_cell = sheet.cell(row=row, column=12)
        if not _is_blank(tjlp_cell.value) and tjlp_cell.number_format != _PERCENT_NUMBER_FORMAT:
            tjlp_cell.number_format = _PERCENT_NUMBER_FORMAT

        selic_cell = sheet.cell(row=row, column=13)
        if (
            not _is_blank(selic_cell.value)
            and selic_cell.number_format != _PERCENT_NUMBER_FORMAT
        ):
            selic_cell.number_format = _PERCENT_NUMBER_FORMAT

        cdi_cell = sheet.cell(row=row, column=14)
        if not _is_blank(cdi_cell.value) and cdi_cell.number_format != _CDI_NUMBER_FORMAT:
            cdi_cell.number_format = _CDI_NUMBER_FORMAT


def _find_last_updated_row(sheet) -> int:
    last_logged = None
    last_date = None
    for row in range(3, sheet.max_row + 1):
        if _coerce_date(sheet.cell(row=row, column=1).value):
            last_date = row
        if sheet.cell(row=row, column=_LOG_COLUMN_INDEX).value:
            last_logged = row
    if last_logged is not None:
        return last_logged
    if last_date is not None:
        return last_date
    raise ValueError("nenhuma linha com data encontrada na planilha")


def _format_date_cell(value: object) -> str:
    cell_date = _coerce_date(value)
    if cell_date:
        return cell_date.strftime("%d/%m/%Y")
    return ""


def _to_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        cleaned = re.sub(r"[^\d,.-]", "", text)
        if "," in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        return Decimal(cleaned)
    raise TypeError(f"valor numerico invalido: {value!r}")


def _format_number_cell(value: object) -> str:
    number = _to_decimal(value)
    if number is None:
        return ""
    formatted = _quantize_4(number)
    return f"{formatted:.4f}".replace(".", ",")


def _format_percent_cell(value: object) -> str:
    number = _to_decimal(value)
    if number is None:
        return ""
    # Compatibilidade: em planilhas novas o valor fica fracionario (0,0919),
    # em planilhas antigas pode aparecer em pontos percentuais (9,19).
    normalized = number if abs(number) > 1 else number * Decimal("100")
    formatted = _quantize_4(normalized)
    return f"{formatted:.4f}".replace(".", ",") + "%"


def _format_cdi_cell(value: object) -> str:
    number = _to_decimal(value)
    if number is None:
        return ""
    formatted = _quantize_10(number)
    return f"{formatted:.10f}".replace(".", ",")


def _format_log_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"OK {value.strftime('%d/%m/%Y %H:%M:%S')}"
    return str(value).strip()


def update_xlsx_usd_brl(
    path: str | Path,
    quote: Quote,
    spread: Decimal = Decimal("0.0020"),
    *,
    target_date: date | None = None,
    overwrite: bool = True,
) -> None:
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)

    def _apply(sheet) -> None:
        collected_date = target_date or _as_local_date(quote.collected_at)
        compra = _quantize_4(quote.value)

        row = _find_or_create_row_by_date(sheet, collected_date)
        _set_cell(
            sheet,
            f"A{row}",
            collected_date,
            number_format=_DATE_NUMBER_FORMAT,
            overwrite=True,
        )
        buy_address = f"B{row}"
        existing_buy = sheet[buy_address].value
        wrote_buy = _set_cell(
            sheet,
            buy_address,
            compra,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )
        buy_for_sale = compra
        if not wrote_buy and not _is_blank(existing_buy):
            try:
                parsed = _to_decimal(existing_buy)
                if parsed is not None:
                    buy_for_sale = _quantize_4(parsed)
            except Exception:
                buy_for_sale = compra

        venda = _quantize_4(buy_for_sale + spread)
        _set_cell(
            sheet,
            f"C{row}",
            venda,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )

    _load_and_save_workbook(target, _apply)


def update_xlsx_dolar_turismo(
    path: str | Path,
    quote: BidAskQuote,
    target_date: date | None = None,
    *,
    overwrite: bool = True,
) -> None:
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)

    def _apply(sheet) -> None:
        use_date = target_date or _as_local_date(quote.collected_at)
        row = _find_or_create_row_by_date(sheet, use_date)

        compra = _quantize_4(quote.buy)
        venda = _quantize_4(quote.sell)

        _set_cell(
            sheet,
            f"F{row}",
            compra,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )
        _set_cell(
            sheet,
            f"G{row}",
            venda,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )

    _load_and_save_workbook(target, _apply)


def update_xlsx_dolar_ptax(
    path: str | Path,
    quote: PtaxQuote,
    target_date: date | None = None,
    *,
    overwrite: bool = True,
) -> None:
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)

    def _apply(sheet) -> None:
        use_date = target_date or _as_local_date(quote.collected_at)
        row = _find_or_create_row_by_date(sheet, use_date)

        compra = _quantize_4(quote.buy)
        venda = _quantize_4(quote.sell)

        _set_cell(
            sheet,
            f"D{row}",
            compra,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )
        _set_cell(
            sheet,
            f"E{row}",
            venda,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )

    _load_and_save_workbook(target, _apply)


def update_xlsx_euro_ptax(
    path: str | Path,
    quote: PtaxQuote,
    target_date: date | None = None,
    *,
    overwrite: bool = True,
) -> None:
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)

    def _apply(sheet) -> None:
        use_date = target_date or _as_local_date(quote.collected_at)
        row = _find_or_create_row_by_date(sheet, use_date)

        compra = _quantize_4(quote.buy)
        venda = _quantize_4(quote.sell)

        _set_cell(
            sheet,
            f"H{row}",
            compra,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )
        _set_cell(
            sheet,
            f"I{row}",
            venda,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )

    _load_and_save_workbook(target, _apply)


def update_xlsx_chf_ptax(
    path: str | Path,
    quote: PtaxQuote,
    target_date: date | None = None,
    *,
    overwrite: bool = True,
) -> None:
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)

    def _apply(sheet) -> None:
        use_date = target_date or _as_local_date(quote.collected_at)
        row = _find_or_create_row_by_date(sheet, use_date)

        compra = _quantize_4(quote.buy)
        venda = _quantize_4(quote.sell)

        _set_cell(
            sheet,
            f"J{row}",
            compra,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )
        _set_cell(
            sheet,
            f"K{row}",
            venda,
            number_format=_QUOTE_NUMBER_FORMAT,
            overwrite=overwrite,
        )

    _load_and_save_workbook(target, _apply)


def update_xlsx_log(
    path: str | Path,
    target_date: date | None = None,
    logged_at: datetime | None = None,
    status: str = "OK",
    detail: str | None = None,
) -> None:
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)

    def _apply(sheet) -> None:
        use_date = target_date or datetime.now(_LOCAL_TZ).date()
        row = _find_or_create_row_by_date(sheet, use_date)

        when = _as_local_datetime(logged_at) if logged_at else datetime.now(_LOCAL_TZ)
        timestamp = when.strftime("%d/%m/%Y %H:%M:%S")
        status_text = (status or "OK").strip()
        if detail:
            detail_text = " ".join(str(detail).split())
            cell_value = f"{status_text} {timestamp} - {detail_text}"
        else:
            cell_value = f"{status_text} {timestamp}"

        _set_cell(sheet, f"{_LOG_COLUMN}{row}", cell_value, overwrite=True)

    _load_and_save_workbook(target, _apply)


def update_xlsx_quotes_and_log(
    path: str | Path,
    *,
    target_date: date,
    usd_brl: Quote | None = None,
    ptax_usd: PtaxQuote | None = None,
    turismo: BidAskQuote | None = None,
    ptax_eur: PtaxQuote | None = None,
    ptax_chf: PtaxQuote | None = None,
    tjlp: Decimal | None = None,
    selic: Decimal | None = None,
    cdi: Decimal | None = None,
    spread: Decimal = Decimal("0.0020"),
    overwrite_quotes: bool = False,
    logged_at: datetime | None = None,
    status: str = "OK",
    detail: str | None = None,
) -> dict[str, tuple[str, ...]]:
    """Atualiza varias cotacoes no XLSX com apenas um save.

    Retorna, por fonte, quais campos foram efetivamente gravados
    (por padrao nao sobrescreve celulas ja preenchidas).
    """
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)

    def _apply(sheet) -> dict[str, tuple[str, ...]]:
        row = _find_or_create_row_by_date(sheet, target_date)
        _set_cell(
            sheet,
            f"A{row}",
            target_date,
            number_format=_DATE_NUMBER_FORMAT,
            overwrite=True,
        )

        written: dict[str, tuple[str, ...]] = {}

        def _append_written(source: str, field: str) -> None:
            existing = list(written.get(source, ()))
            existing.append(field)
            written[source] = tuple(existing)

        if usd_brl:
            compra = _quantize_4(usd_brl.value)

            buy_address = f"B{row}"
            existing_buy = sheet[buy_address].value
            wrote_buy = _set_cell(
                sheet,
                buy_address,
                compra,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            )
            if wrote_buy:
                _append_written("usd_brl", "compra")

            buy_for_sale = compra
            if not wrote_buy and not _is_blank(existing_buy):
                try:
                    parsed = _to_decimal(existing_buy)
                    if parsed is not None:
                        buy_for_sale = _quantize_4(parsed)
                except Exception:
                    buy_for_sale = compra

            venda = _quantize_4(buy_for_sale + spread)
            if _set_cell(
                sheet,
                f"C{row}",
                venda,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("usd_brl", "venda")

        if ptax_usd:
            compra = _quantize_4(ptax_usd.buy)
            venda = _quantize_4(ptax_usd.sell)
            if _set_cell(
                sheet,
                f"D{row}",
                compra,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("ptax_usd", "compra")
            if _set_cell(
                sheet,
                f"E{row}",
                venda,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("ptax_usd", "venda")

        if turismo:
            compra = _quantize_4(turismo.buy)
            venda = _quantize_4(turismo.sell)
            if _set_cell(
                sheet,
                f"F{row}",
                compra,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("turismo", "compra")
            if _set_cell(
                sheet,
                f"G{row}",
                venda,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("turismo", "venda")

        if ptax_eur:
            compra = _quantize_4(ptax_eur.buy)
            venda = _quantize_4(ptax_eur.sell)
            if _set_cell(
                sheet,
                f"H{row}",
                compra,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("ptax_eur", "compra")
            if _set_cell(
                sheet,
                f"I{row}",
                venda,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("ptax_eur", "venda")

        if ptax_chf:
            compra = _quantize_4(ptax_chf.buy)
            venda = _quantize_4(ptax_chf.sell)
            if _set_cell(
                sheet,
                f"J{row}",
                compra,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("ptax_chf", "compra")
            if _set_cell(
                sheet,
                f"K{row}",
                venda,
                number_format=_QUOTE_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("ptax_chf", "venda")

        tjlp_percent = _to_decimal(tjlp) if tjlp is not None else None
        if tjlp_percent is not None:
            tjlp_fraction = _quantize_4(tjlp_percent / Decimal("100"))
            if _set_cell(
                sheet,
                f"{_TJLP_COLUMN}{row}",
                tjlp_fraction,
                number_format=_PERCENT_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("tjlp", "valor")

        selic_percent = _to_decimal(selic) if selic is not None else None
        if selic_percent is not None:
            selic_fraction = _quantize_4(selic_percent / Decimal("100"))
            if _set_cell(
                sheet,
                f"{_SELIC_COLUMN}{row}",
                selic_fraction,
                number_format=_PERCENT_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("selic", "selic")

        cdi_percent = _to_decimal(cdi) if cdi is not None else None
        if cdi_percent is not None:
            cdi_value = _quantize_10(cdi_percent)
            if _set_cell(
                sheet,
                f"{_CDI_COLUMN}{row}",
                cdi_value,
                number_format=_CDI_NUMBER_FORMAT,
                overwrite=overwrite_quotes,
            ):
                _append_written("selic", "cdi")

        # Regras do cliente: se nao houver novo valor de juros no dia, repete o ultimo.
        if _repeat_previous_value_if_blank(
            sheet,
            row,
            12,
            number_format=_PERCENT_NUMBER_FORMAT,
        ):
            _append_written("tjlp", "valor_repetido")
        if _repeat_previous_value_if_blank(
            sheet,
            row,
            13,
            number_format=_PERCENT_NUMBER_FORMAT,
        ):
            _append_written("selic", "selic_repetido")
        if _repeat_previous_value_if_blank(
            sheet,
            row,
            14,
            number_format=_CDI_NUMBER_FORMAT,
        ):
            _append_written("selic", "cdi_repetido")

        when = _as_local_datetime(logged_at) if logged_at else datetime.now(_LOCAL_TZ)
        timestamp = when.strftime("%d/%m/%Y %H:%M:%S")
        status_text = (status or "OK").strip()
        if detail:
            detail_text = " ".join(str(detail).split())
            cell_value = f"{status_text} {timestamp} - {detail_text}"
        else:
            cell_value = f"{status_text} {timestamp}"

        _set_cell(sheet, f"{_LOG_COLUMN}{row}", cell_value, overwrite=True)
        return written

    return _load_and_save_workbook(target, _apply)


def normalize_xlsx_layout(path: str | Path) -> None:
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)

    def _apply(_sheet) -> None:
        return None

    _load_and_save_workbook(target, _apply)


def update_csv_from_xlsx(
    xlsx_path: str | Path,
    csv_path: str | Path,
) -> None:
    source = Path(xlsx_path)
    if not source.exists():
        raise FileNotFoundError(source)

    workbook = load_workbook(source, data_only=True)
    try:
        sheet = workbook.active
        _ensure_layout(sheet)
        row = _find_last_updated_row(sheet)
        row_values = [sheet.cell(row=row, column=col).value for col in range(1, 16)]
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()

    date_value = _format_date_cell(row_values[0])
    if not date_value:
        raise ValueError("data da linha nao encontrada para atualizar o CSV")

    data_row = [date_value]
    for index in range(1, 11):
        data_row.append(_format_number_cell(row_values[index]))
    data_row.append(_format_percent_cell(row_values[11]))
    data_row.append(_format_percent_cell(row_values[12]))
    data_row.append(_format_cdi_cell(row_values[13]))
    data_row.append(_format_log_cell(row_values[14]))

    target = Path(csv_path)
    existing_rows: list[list[str]] = []
    if target.exists():
        for encoding in ("utf-8", "latin-1"):
            try:
                with target.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.reader(handle, delimiter=";")
                    existing_rows = list(reader)
                break
            except UnicodeDecodeError:
                existing_rows = []
                continue

    data_rows = [
        row_values
        for row_values in existing_rows
        if row_values and _DATE_PATTERN.match(row_values[0])
    ]
    replaced = False
    new_data_rows: list[list[str]] = []
    for row_values in data_rows:
        if row_values[0] == date_value:
            new_data_rows.append(data_row)
            replaced = True
        else:
            new_data_rows.append(row_values)
    if not replaced:
        new_data_rows.append(data_row)

    rows = [_DEFAULT_CSV_HEADER] + new_data_rows

    with target.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerows(rows)
