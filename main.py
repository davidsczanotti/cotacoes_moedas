from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
import os
from pathlib import Path
import shutil
import sys
import time
from typing import Callable

from openpyxl import load_workbook

from cotacoes_moedas import (
    calculate_cdi_daily_percent,
    fetch_chf_ptax,
    fetch_dolar_ptax,
    fetch_dolar_turismo,
    fetch_euro_ptax,
    fetch_selic,
    fetch_tjlp,
    fetch_usd_brl,
    normalize_xlsx_layout,
    update_csv_from_xlsx,
    update_xlsx_quotes_and_log,
)
from cotacoes_moedas.network_sync import (
    copiar_pasta_para_rede,
    parse_network_dirs,
)
from cotacoes_moedas.network_copy import try_to_unc
from cotacoes_moedas.redaction import redact_secrets

_USD_SPREAD = Decimal("0.0020")
_LOCAL_TZ = datetime.now().astimezone().tzinfo or timezone.utc
_DEFAULT_NETWORK_COPY_DIR = r"X:\TEMP\_Publico;\\192.168.21.25\users\TEMP\_Publico"
_DEFAULT_NETWORK_DEST_FOLDER = "cotacoes"
_MORNING_QUOTES_CUTOFF_HM = (8, 30)
_PTAX_AVAILABLE_FROM_HM = (13, 10)
_SOURCE_REQUIRED_COLUMNS: dict[str, tuple[str, ...]] = {
    "usd_brl": ("B", "C"),
    "ptax_usd": ("D", "E"),
    "turismo": ("F", "G"),
    "ptax_eur": ("H", "I"),
    "ptax_chf": ("J", "K"),
    "tjlp": ("L",),
    "selic": ("M", "N"),
}
_SOURCE_LABELS: dict[str, str] = {
    "usd_brl": "USD/BRL (Investing)",
    "ptax_usd": "PTAX USD",
    "ptax_eur": "PTAX EUR",
    "ptax_chf": "PTAX CHF",
    "turismo": "Dolar Turismo (Valor)",
    "tjlp": "TJLP (BNDES)",
    "selic": "SELIC (BCB)",
}


@dataclass(frozen=True)
class FetchSpec:
    key: str
    label: str
    fetch_fn: Callable[[], object]


@dataclass
class FetchOutcome:
    label: str
    value: object | None
    error: str | None
    elapsed_s: float
    skipped: bool = False
    skip_reason: str | None = None


def _now_local() -> datetime:
    return datetime.now(_LOCAL_TZ)


def _log(message: str) -> None:
    timestamp = _now_local().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}", flush=True)


def _log_stage(step: int, total: int, message: str) -> None:
    _log(f"Etapa {step}/{total} - {message}")


def _format_duration(seconds: float) -> str:
    total_seconds = int(round(seconds))
    minutes, remainder = divmod(total_seconds, 60)
    return f"{minutes:02d}:{remainder:02d}"


def _resolve_base_dir() -> Path:
    return Path(sys.argv[0]).resolve().parent


def _configure_playwright(base_dir: Path) -> None:
    if os.environ.get("PLAYWRIGHT_BROWSERS_PATH"):
        return
    browsers_path = base_dir / "ms-playwright"
    if browsers_path.exists():
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(browsers_path)
        _log(f"Playwright browsers: {browsers_path}")


def _error_detail(label: str, exc: Exception) -> str:
    message = " ".join(str(exc).split())
    message = redact_secrets(message)
    if message:
        return f"{label}: {exc.__class__.__name__} {message}"
    return f"{label}: {exc.__class__.__name__}"


def _run_fetch(
    label: str,
    fetch_fn: Callable[[], object],
) -> tuple[object | None, str | None, float]:
    _log(f"Coletando {label}...")
    start = time.monotonic()
    try:
        result = fetch_fn()
    except Exception as exc:
        detail = _error_detail(label, exc)
        _log(f"Falha em {label}. Valores nao atualizados. Detalhe: {detail}")
        return None, detail, time.monotonic() - start
    elapsed = time.monotonic() - start
    _log(f"{label} OK em {elapsed:.1f}s")
    return result, None, elapsed


def _run_fetches(fetch_specs: list[FetchSpec]) -> dict[str, FetchOutcome]:
    outcomes: dict[str, FetchOutcome] = {}
    if not fetch_specs:
        return outcomes

    max_workers = len(fetch_specs)
    env_max_workers = os.environ.get("COTACOES_MAX_WORKERS")
    if env_max_workers:
        try:
            max_workers = max(1, min(max_workers, int(env_max_workers)))
        except ValueError:
            _log(
                "Aviso: COTACOES_MAX_WORKERS invalido "
                f"({env_max_workers!r}). Usando {max_workers}."
            )

    if max_workers <= 1 or len(fetch_specs) == 1:
        _log(f"Coleta: {len(fetch_specs)} fonte(s) em modo sequencial.")
        for spec in fetch_specs:
            value, error, elapsed = _run_fetch(spec.label, spec.fetch_fn)
            outcomes[spec.key] = FetchOutcome(
                label=spec.label,
                value=value,
                error=error,
                elapsed_s=elapsed,
            )
        return outcomes

    limit_note = (
        f" (limitado por COTACOES_MAX_WORKERS={env_max_workers})"
        if env_max_workers
        else ""
    )
    _log(
        "Coleta: "
        f"{len(fetch_specs)} fonte(s) em paralelo ({max_workers} workers){limit_note}."
    )
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(_run_fetch, spec.label, spec.fetch_fn): spec
            for spec in fetch_specs
        }
        for future in as_completed(futures):
            spec = futures[future]
            try:
                value, error, elapsed = future.result()
            except Exception as exc:
                detail = _error_detail(spec.label, exc)
                _log(
                    f"Falha em {spec.label}. Valores nao atualizados. "
                    f"Detalhe: {detail}"
                )
                value, error, elapsed = None, detail, 0.0
            outcomes[spec.key] = FetchOutcome(
                label=spec.label,
                value=value,
                error=error,
                elapsed_s=elapsed,
            )
    return outcomes


def _collect_errors(outcomes: dict[str, FetchOutcome]) -> list[str]:
    errors: list[str] = []
    for outcome in outcomes.values():
        if outcome.error:
            errors.append(outcome.error)
    return errors


def _log_fetch_summary(outcomes: dict[str, FetchOutcome]) -> None:
    total = len(outcomes)
    skipped = sum(1 for outcome in outcomes.values() if outcome.skipped)
    attempted = total - skipped
    ok_count = sum(
        1
        for outcome in outcomes.values()
        if not outcome.skipped and not outcome.error
    )
    fail_count = sum(
        1
        for outcome in outcomes.values()
        if not outcome.skipped and outcome.error
    )
    _log(
        "Coleta concluida: "
        f"{ok_count}/{attempted} fontes OK, "
        f"{fail_count} com erro, "
        f"{skipped} puladas."
    )


def _log_fetch_plan(
    selected_specs: list[FetchSpec],
    outcomes: dict[str, FetchOutcome],
) -> None:
    if selected_specs:
        labels = "; ".join(spec.label for spec in selected_specs)
        _log(f"Fontes selecionadas para coleta ({len(selected_specs)}): {labels}")
    else:
        _log("Nenhuma fonte selecionada para coleta agora.")

    for key in (
        "usd_brl",
        "turismo",
        "ptax_usd",
        "ptax_eur",
        "ptax_chf",
        "tjlp",
        "selic",
    ):
        outcome = outcomes.get(key)
        if outcome and outcome.skipped:
            _log(f"{outcome.label}: pulado ({outcome.skip_reason})")


def _hm(value: datetime) -> tuple[int, int]:
    return value.hour, value.minute


def _coerce_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _find_row_by_date(sheet, target_date: date) -> int | None:
    for row in range(3, sheet.max_row + 1):
        cell_date = _coerce_date(sheet.cell(row=row, column=1).value)
        if cell_date == target_date:
            return row
    return None


def _is_source_filled(sheet, row: int, columns: tuple[str, ...]) -> bool:
    for col in columns:
        value = sheet[f"{col}{row}"].value
        if value is None:
            return False
        if isinstance(value, str) and not value.strip():
            return False
    return True


def _read_filled_sources(planilha_path: Path, target_date: date) -> dict[str, bool]:
    workbook = load_workbook(planilha_path, data_only=True)
    try:
        sheet = workbook.active
        row = _find_row_by_date(sheet, target_date)
        filled: dict[str, bool] = {}
        for key, columns in _SOURCE_REQUIRED_COLUMNS.items():
            filled[key] = (
                row is not None and _is_source_filled(sheet, row, columns)
            )
        return filled
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def _skip_outcome(key: str, reason: str) -> FetchOutcome:
    return FetchOutcome(
        label=_SOURCE_LABELS[key],
        value=None,
        error=None,
        elapsed_s=0.0,
        skipped=True,
        skip_reason=reason,
    )


def _empty_filled_sources() -> dict[str, bool]:
    return {key: False for key in _SOURCE_REQUIRED_COLUMNS}


def _network_planilha_candidates(
    network_dirs: list[str],
    *,
    network_dest_folder: str,
) -> list[Path]:
    candidates: list[Path] = []
    seen: set[str] = set()

    for base_dir in network_dirs:
        raw_base = (base_dir or "").strip()
        if not raw_base:
            continue

        bases = [raw_base]
        unc_base, _ = try_to_unc(raw_base)
        if unc_base and unc_base != raw_base:
            bases.append(unc_base)

        for base in bases:
            candidate = (
                Path(base) / network_dest_folder / "planilhas" / "cotacoes.xlsx"
            )
            key = str(candidate).lower()
            if key in seen:
                continue
            seen.add(key)
            candidates.append(candidate)

    return candidates


def _select_reference_planilha_path(
    local_planilha_path: Path,
    *,
    network_dirs: list[str],
    network_dest_folder: str,
) -> Path | None:
    if not network_dirs:
        return local_planilha_path

    candidates = _network_planilha_candidates(
        network_dirs,
        network_dest_folder=network_dest_folder,
    )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    if candidates:
        return candidates[0]
    return None


def _same_path(left: Path, right: Path) -> bool:
    return str(left).strip().lower() == str(right).strip().lower()


def _sync_local_planilhas_from_reference(
    reference_planilha_path: Path,
    *,
    local_planilha_path: Path,
    local_csv_path: Path,
) -> bool:
    if _same_path(reference_planilha_path, local_planilha_path):
        return True

    if not reference_planilha_path.exists():
        _log(
            "ERRO: Planilha de referencia nao encontrada para sincronizacao local: "
            f"{reference_planilha_path}"
        )
        return False

    if local_planilha_path.exists():
        try:
            reference_mtime = reference_planilha_path.stat().st_mtime
            local_mtime = local_planilha_path.stat().st_mtime
        except OSError as exc:
            detail = redact_secrets(str(exc))
            _log(
                "Aviso: Nao foi possivel comparar datas de modificacao entre "
                "planilha local e referencia da rede: "
                f"{exc.__class__.__name__} {detail}. "
                "Mantendo base local."
            )
            return True
        if reference_mtime <= local_mtime:
            _log(
                "Sincronizacao local: referencia da rede nao e mais nova que o "
                "arquivo local. Mantendo base local para evitar sobrescrita."
            )
            return True

    _log(
        "Sincronizando arquivos locais a partir da planilha de referencia da rede..."
    )
    try:
        local_planilha_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(reference_planilha_path, local_planilha_path)
        reference_csv_path = reference_planilha_path.with_name("cotacoes.csv")
        if reference_csv_path.exists():
            shutil.copy2(reference_csv_path, local_csv_path)
    except Exception as exc:
        detail = redact_secrets(str(exc))
        _log(
            "ERRO: Falha ao sincronizar base local com a referencia da rede: "
            f"{exc.__class__.__name__} {detail}"
        )
        return False

    _log("Sincronizacao local concluida.")
    return True


def _validate_planilha_row_consistency(
    planilha_path: Path,
    *,
    target_date: date,
    outcomes: dict[str, FetchOutcome],
) -> list[str]:
    if not planilha_path.exists():
        return [f"planilha nao encontrada: {planilha_path}"]

    workbook = load_workbook(planilha_path, data_only=True)
    try:
        sheet = workbook.active
        row = _find_row_by_date(sheet, target_date)
        if row is None:
            return [
                "linha da data nao encontrada: "
                f"{target_date.strftime('%d/%m/%Y')} em {planilha_path}"
            ]

        issues: list[str] = []
        for key, columns in _SOURCE_REQUIRED_COLUMNS.items():
            outcome = outcomes.get(key)
            if outcome is None:
                continue
            should_be_filled = (
                outcome.skip_reason == "ja preenchido na data de hoje"
                or outcome.value is not None
            )
            if should_be_filled and not _is_source_filled(sheet, row, columns):
                issues.append(
                    f"{_SOURCE_LABELS[key]}: colunas esperadas "
                    f"{'/'.join(columns)} vazias na linha {row}"
                )
        return issues
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def _select_fetches(
    now: datetime,
    planilha_path: Path,
    *,
    reference_planilha_path: Path | None = None,
) -> tuple[list[FetchSpec], dict[str, FetchOutcome]]:
    today = now.date()
    source_for_validation = reference_planilha_path or planilha_path
    if source_for_validation.exists():
        filled = _read_filled_sources(source_for_validation, today)
    else:
        filled = _empty_filled_sources()
    allow_morning_quotes = _hm(now) <= _MORNING_QUOTES_CUTOFF_HM
    allow_ptax = _hm(now) >= _PTAX_AVAILABLE_FROM_HM

    all_specs = {
        "usd_brl": FetchSpec(
            key="usd_brl",
            label=_SOURCE_LABELS["usd_brl"],
            fetch_fn=fetch_usd_brl,
        ),
        "ptax_usd": FetchSpec(
            key="ptax_usd",
            label=_SOURCE_LABELS["ptax_usd"],
            fetch_fn=fetch_dolar_ptax,
        ),
        "ptax_eur": FetchSpec(
            key="ptax_eur",
            label=_SOURCE_LABELS["ptax_eur"],
            fetch_fn=fetch_euro_ptax,
        ),
        "ptax_chf": FetchSpec(
            key="ptax_chf",
            label=_SOURCE_LABELS["ptax_chf"],
            fetch_fn=fetch_chf_ptax,
        ),
        "turismo": FetchSpec(
            key="turismo",
            label=_SOURCE_LABELS["turismo"],
            fetch_fn=fetch_dolar_turismo,
        ),
        "tjlp": FetchSpec(
            key="tjlp",
            label=_SOURCE_LABELS["tjlp"],
            fetch_fn=fetch_tjlp,
        ),
        "selic": FetchSpec(
            key="selic",
            label=_SOURCE_LABELS["selic"],
            fetch_fn=fetch_selic,
        ),
    }

    selected: list[FetchSpec] = []
    outcomes: dict[str, FetchOutcome] = {}

    for key in ("usd_brl", "turismo"):
        if not allow_morning_quotes:
            outcomes[key] = _skip_outcome(key, "fora do horario (apos 08:30)")
        elif filled.get(key, False):
            outcomes[key] = _skip_outcome(key, "ja preenchido na data de hoje")
        else:
            selected.append(all_specs[key])

    for key in ("ptax_usd", "ptax_eur", "ptax_chf"):
        if not allow_ptax:
            outcomes[key] = _skip_outcome(key, "fora do horario (antes de 13:10)")
        elif filled.get(key, False):
            outcomes[key] = _skip_outcome(key, "ja preenchido na data de hoje")
        else:
            selected.append(all_specs[key])

    for key in ("tjlp", "selic"):
        if not allow_morning_quotes:
            outcomes[key] = _skip_outcome(key, "fora do horario (apos 08:30)")
        elif filled.get(key, False):
            outcomes[key] = _skip_outcome(key, "ja preenchido na data de hoje")
        else:
            selected.append(all_specs[key])

    return selected, outcomes


def _update_planilha(
    planilha_path: Path,
    target_date: date,
    outcomes: dict[str, FetchOutcome],
    errors: list[str],
) -> dict[str, tuple[str, ...]]:
    _log(f"Atualizando planilha: {planilha_path} (gravacao unica)")

    all_errors = list(errors)
    tjlp_quote = outcomes["tjlp"].value
    selic_quote = outcomes["selic"].value
    selic_percent = selic_quote.value if selic_quote else None
    cdi_daily_percent = None
    if selic_percent is not None:
        try:
            cdi_daily_percent = calculate_cdi_daily_percent(selic_percent)
        except Exception as exc:
            all_errors.append(_error_detail("CDI", exc))

    errors[:] = all_errors

    status = "ERRO" if all_errors else "OK"
    detail = " | ".join(all_errors) if all_errors else None

    written = update_xlsx_quotes_and_log(
        planilha_path,
        target_date=target_date,
        usd_brl=outcomes["usd_brl"].value,
        ptax_usd=outcomes["ptax_usd"].value,
        ptax_eur=outcomes["ptax_eur"].value,
        ptax_chf=outcomes["ptax_chf"].value,
        turismo=outcomes["turismo"].value,
        tjlp=tjlp_quote.value if tjlp_quote else None,
        selic=selic_percent,
        cdi=cdi_daily_percent,
        spread=_USD_SPREAD,
        overwrite_quotes=False,
        logged_at=_now_local(),
        status=status,
        detail=detail,
    )

    def _describe_fields(fields: tuple[str, ...]) -> str:
        if not fields:
            return "nao gravou (ja preenchido na planilha)"
        if fields == ("compra", "venda"):
            return "gravou compra e venda"
        descriptions = {
            "valor": "valor",
            "valor_repetido": "ultimo valor",
            "selic": "SELIC",
            "selic_repetido": "SELIC (ultimo valor)",
            "cdi": "CDI",
            "cdi_repetido": "CDI (ultimo valor)",
        }
        return "gravou " + " e ".join(descriptions.get(field, field) for field in fields)

    for key in (
        "usd_brl",
        "turismo",
        "ptax_usd",
        "ptax_eur",
        "ptax_chf",
        "tjlp",
        "selic",
    ):
        outcome = outcomes[key]
        fields = written.get(key, ())
        if outcome.skipped:
            _log(f"{outcome.label} pulado: {outcome.skip_reason}")
            continue
        if outcome.value is None and not fields:
            _log(f"{outcome.label}: sem dados; planilha nao atualizada para esta fonte.")
            continue
        _log(f"{outcome.label}: {_describe_fields(fields)}.")

    has_quotes = any(
        outcome.value is not None for outcome in outcomes.values() if not outcome.skipped
    )
    wrote_any = any(fields for fields in written.values())
    if has_quotes and not wrote_any:
        _log(
            "Nenhuma cotacao foi gravada (valores ja estavam preenchidos). "
            "Apenas o log foi atualizado."
        )

    return written


def _log_quote_summary(outcomes: dict[str, FetchOutcome]) -> None:
    usd = outcomes["usd_brl"]
    quote = usd.value
    if quote:
        _log(
            "USD/BRL: "
            f"{quote.value} ({quote.value_raw}) em {quote.collected_at.astimezone(_LOCAL_TZ)}"
        )
    elif usd.skipped:
        _log(f"USD/BRL: pulado ({usd.skip_reason})")
    else:
        _log("USD/BRL: sem dados")

    usd_ptax_outcome = outcomes["ptax_usd"]
    ptax = usd_ptax_outcome.value
    if ptax:
        _log(
            "PTAX USD: "
            f"{ptax.buy} / {ptax.sell} em {ptax.collected_at.astimezone(_LOCAL_TZ)}"
        )
    elif usd_ptax_outcome.skipped:
        _log(f"PTAX USD: pulado ({usd_ptax_outcome.skip_reason})")
    else:
        _log("PTAX USD: sem dados")

    eur_ptax_outcome = outcomes["ptax_eur"]
    euro = eur_ptax_outcome.value
    if euro:
        _log(
            "PTAX EUR: "
            f"{euro.buy} / {euro.sell} em {euro.collected_at.astimezone(_LOCAL_TZ)}"
        )
    elif eur_ptax_outcome.skipped:
        _log(f"PTAX EUR: pulado ({eur_ptax_outcome.skip_reason})")
    else:
        _log("PTAX EUR: sem dados")

    chf_ptax_outcome = outcomes["ptax_chf"]
    chf = chf_ptax_outcome.value
    if chf:
        _log(
            "PTAX CHF: "
            f"{chf.buy} / {chf.sell} em {chf.collected_at.astimezone(_LOCAL_TZ)}"
        )
    elif chf_ptax_outcome.skipped:
        _log(f"PTAX CHF: pulado ({chf_ptax_outcome.skip_reason})")
    else:
        _log("PTAX CHF: sem dados")

    turismo_outcome = outcomes["turismo"]
    turismo = turismo_outcome.value
    if turismo:
        _log(
            "Dolar Turismo: "
            f"{turismo.buy} / {turismo.sell} em {turismo.collected_at.astimezone(_LOCAL_TZ)}"
        )
    elif turismo_outcome.skipped:
        _log(f"Dolar Turismo: pulado ({turismo_outcome.skip_reason})")
    else:
        _log("Dolar Turismo: sem dados")

    tjlp_outcome = outcomes["tjlp"]
    tjlp = tjlp_outcome.value
    if tjlp:
        _log(
            "TJLP: "
            f"{tjlp.value:.4f}% em {tjlp.collected_at.astimezone(_LOCAL_TZ)}"
        )
    elif tjlp_outcome.skipped:
        _log(f"TJLP: pulado ({tjlp_outcome.skip_reason})")
    else:
        _log("TJLP: sem dados")

    selic_outcome = outcomes["selic"]
    selic = selic_outcome.value
    if selic:
        details = (
            f"{selic.value:.4f}%"
            f" (referencia {selic.reference_date.strftime('%d/%m/%Y')})"
            if selic.reference_date
            else f"{selic.value:.4f}%"
        )
        _log(
            "SELIC: "
            f"{details} em {selic.collected_at.astimezone(_LOCAL_TZ)}"
        )
        try:
            cdi = calculate_cdi_daily_percent(selic.value)
            _log(f"CDI (calculado): {cdi:.10f}")
        except Exception as exc:
            _log(f"CDI (calculado): erro ({exc.__class__.__name__}: {exc})")
    elif selic_outcome.skipped:
        _log(f"SELIC: pulado ({selic_outcome.skip_reason})")
    else:
        _log("SELIC: sem dados")


def _copy_planilhas_to_network(
    planilhas_dir: Path,
    network_dirs: list[str],
    *,
    network_dest_folder: str,
) -> Path | None:
    if not planilhas_dir.exists() or not planilhas_dir.is_dir():
        _log(f"Pasta de planilhas nao encontrada: {planilhas_dir}")
        return None

    _log("Copiando pasta de planilhas para a rede...")
    destination_dir, unc_error, copy_error = copiar_pasta_para_rede(
        planilhas_dir,
        network_dirs,
        nome_pasta_destino=network_dest_folder,
    )
    if not destination_dir:
        if copy_error:
            detail = redact_secrets(str(copy_error))
            _log(
                "Copia em rede falhou: "
                f"{copy_error.__class__.__name__} {detail}"
            )
        else:
            _log(
                "Copia em rede: nenhum destino valido informado: "
                + "; ".join(network_dirs)
            )
        return None

    if unc_error:
        detail = redact_secrets(str(unc_error))
        _log(
            "Aviso: Nao foi possivel converter para UNC, usando caminho original: "
            f"{detail}"
        )

    _log(f"Pasta '{planilhas_dir.name}' copiada para: {destination_dir}")
    return destination_dir


def main() -> int:
    process_start = time.monotonic()
    total_steps = 6

    try:
        _log("Inicio da coleta de cotacoes.")
        base_dir = _resolve_base_dir()

        _log_stage(1, total_steps, "Preparando ambiente e configuracoes.")
        _log(f"Diretorio base: {base_dir}")
        _configure_playwright(base_dir)
        network_copy_dirs = parse_network_dirs(
            os.environ.get("COTACOES_NETWORK_DIR") or _DEFAULT_NETWORK_COPY_DIR
        )
        planilha_path = base_dir / "planilhas" / "cotacoes.xlsx"
        csv_path = base_dir / "planilhas" / "cotacoes.csv"
        _log(f"Planilha principal: {planilha_path}")
        _log(f"Arquivo CSV: {csv_path}")
        if network_copy_dirs:
            _log("Destinos de copia em rede: " + "; ".join(network_copy_dirs))
        else:
            _log("Destinos de copia em rede: nenhum")
        network_dest_folder = (
            os.environ.get("COTACOES_NETWORK_DEST_FOLDER")
            or _DEFAULT_NETWORK_DEST_FOLDER
        ).strip()
        if not network_dest_folder:
            network_dest_folder = _DEFAULT_NETWORK_DEST_FOLDER
        _log(f"Pasta de destino na rede: {network_dest_folder}")

        _log_stage(2, total_steps, "Coletando cotacoes das fontes.")
        now = _now_local()
        now_hm = _hm(now)
        _log(
            "Horario local atual: "
            f"{now.strftime('%H:%M')} "
            f"(Investing/Valor ate {(_MORNING_QUOTES_CUTOFF_HM[0]):02d}:{(_MORNING_QUOTES_CUTOFF_HM[1]):02d}; "
            f"PTAX apos {(_PTAX_AVAILABLE_FROM_HM[0]):02d}:{(_PTAX_AVAILABLE_FROM_HM[1]):02d}; "
            "TJLP/SELIC ate 08:30)."
        )
        if now_hm > _MORNING_QUOTES_CUTOFF_HM and now_hm < _PTAX_AVAILABLE_FROM_HM:
            _log(
                "Fora da janela de coleta no momento "
                "(apos 08:30 e antes de 13:10). "
                "Sem coleta agora, mas validacao/bootstrap de rede continua."
            )
        _log(
            "Validando planilha para decidir quais fontes coletar "
            "(nao sobrescreve valores ja preenchidos no dia)."
        )
        reference_planilha_path = _select_reference_planilha_path(
            planilha_path,
            network_dirs=network_copy_dirs,
            network_dest_folder=network_dest_folder,
        )
        if reference_planilha_path is None:
            _log(
                "ERRO: Nao foi possivel resolver caminho de planilha para validacao "
                "na rede. Execucao abortada."
            )
            duration = _format_duration(time.monotonic() - process_start)
            _log(f"Processo abortado em {duration} (minutos:segundos).")
            return 1
        elif reference_planilha_path.exists():
            source_label = "rede" if reference_planilha_path != planilha_path else "local"
            _log(
                "Planilha de validacao de preenchimento "
                f"({source_label}): {reference_planilha_path}"
            )
        else:
            _log(
                "Planilha de validacao na rede nao encontrada: "
                f"{reference_planilha_path}. "
                "Copiando planilhas locais para inicializar o destino."
            )
            try:
                normalize_xlsx_layout(planilha_path)
            except Exception as exc:
                detail = _error_detail("Formatacao local", exc)
                _log(f"ERRO ao preparar planilha local para bootstrap: {detail}")
                duration = _format_duration(time.monotonic() - process_start)
                _log(f"Processo abortado em {duration} (minutos:segundos).")
                return 1
            bootstrapped_dir = _copy_planilhas_to_network(
                base_dir / "planilhas",
                network_copy_dirs,
                network_dest_folder=network_dest_folder,
            )
            if not bootstrapped_dir:
                _log("ERRO: Falha ao inicializar planilha de referencia na rede.")
                duration = _format_duration(time.monotonic() - process_start)
                _log(f"Processo abortado em {duration} (minutos:segundos).")
                return 1

            bootstrapped_reference = bootstrapped_dir / "cotacoes.xlsx"
            if not bootstrapped_reference.exists():
                _log(
                    "ERRO: Inicializacao concluida, mas planilha de referencia "
                    f"nao foi encontrada em {bootstrapped_reference}."
                )
                duration = _format_duration(time.monotonic() - process_start)
                _log(f"Processo abortado em {duration} (minutos:segundos).")
                return 1

            reference_planilha_path = bootstrapped_reference
            _log(
                "Planilha de validacao de preenchimento (rede/bootstrap): "
                f"{reference_planilha_path}"
            )

        if not _sync_local_planilhas_from_reference(
            reference_planilha_path,
            local_planilha_path=planilha_path,
            local_csv_path=csv_path,
        ):
            duration = _format_duration(time.monotonic() - process_start)
            _log(f"Processo abortado em {duration} (minutos:segundos).")
            return 1

        selected_specs, outcomes = _select_fetches(
            now,
            planilha_path,
            reference_planilha_path=reference_planilha_path,
        )
        _log_fetch_plan(selected_specs, outcomes)
        if not selected_specs:
            _log(
                "Nenhuma fonte elegivel para atualizar agora "
                "(fora da janela de horario ou ja preenchida). "
                "Aplicando apenas normalizacao e sincronizacao final."
            )

            _log_stage(3, total_steps, "Normalizando layout da planilha Excel.")
            try:
                normalize_xlsx_layout(planilha_path)
            except Exception as exc:
                detail = _error_detail("Formatacao local", exc)
                _log(f"ERRO ao normalizar planilha local: {detail}")
                duration = _format_duration(time.monotonic() - process_start)
                _log(f"Processo abortado em {duration} (minutos:segundos).")
                return 1

            _log_stage(4, total_steps, "CSV mantido (sem novas cotacoes).")
            _log("CSV nao atualizado porque nao houve coleta de novas fontes.")

            _log_stage(5, total_steps, "Resumo das cotacoes coletadas.")
            _log("Coleta nao executada nesta janela.")

            _log_stage(
                6,
                total_steps,
                f"Validando pasta '{network_dest_folder}' na rede e copiando planilhas.",
            )
            if not network_copy_dirs:
                _log("Copia em rede desabilitada.")
            else:
                destination_planilhas_dir = _copy_planilhas_to_network(
                    base_dir / "planilhas",
                    network_copy_dirs,
                    network_dest_folder=network_dest_folder,
                )
                if destination_planilhas_dir:
                    _log("Sincronizacao final na rede concluida.")
            duration = _format_duration(time.monotonic() - process_start)
            _log(f"Processo finalizado em {duration} (minutos:segundos).")
            return 0

        outcomes.update(_run_fetches(selected_specs))
        errors = _collect_errors(outcomes)
        _log_fetch_summary(outcomes)

        _log_stage(3, total_steps, "Atualizando planilha Excel.")
        _update_planilha(planilha_path, now.date(), outcomes, errors)
        local_validation_issues = _validate_planilha_row_consistency(
            planilha_path,
            target_date=now.date(),
            outcomes=outcomes,
        )
        if local_validation_issues:
            _log("ERRO: Validacao local apos atualizacao encontrou inconsistencias:")
            for issue in local_validation_issues:
                _log(f"- {issue}")
            duration = _format_duration(time.monotonic() - process_start)
            _log(f"Processo abortado em {duration} (minutos:segundos).")
            return 1

        _log_stage(4, total_steps, "Atualizando CSV.")
        update_csv_from_xlsx(planilha_path, csv_path)

        _log_stage(5, total_steps, "Resumo das cotacoes coletadas.")
        _log_quote_summary(outcomes)
        if errors:
            _log(f"Falhas na coleta: {len(errors)}. Consulte o log da planilha.")
        else:
            _log("Coleta sem falhas.")

        _log_stage(
            6,
            total_steps,
            f"Validando pasta '{network_dest_folder}' na rede e copiando planilhas.",
        )
        if not network_copy_dirs:
            _log("Copia em rede desabilitada.")
        else:
            destination_planilhas_dir = _copy_planilhas_to_network(
                base_dir / "planilhas",
                network_copy_dirs,
                network_dest_folder=network_dest_folder,
            )
            if destination_planilhas_dir:
                network_validation_issues = _validate_planilha_row_consistency(
                    destination_planilhas_dir / "cotacoes.xlsx",
                    target_date=now.date(),
                    outcomes=outcomes,
                )
                if network_validation_issues:
                    _log(
                        "ERRO: Validacao final na rede encontrou inconsistencias:"
                    )
                    for issue in network_validation_issues:
                        _log(f"- {issue}")
                    duration = _format_duration(time.monotonic() - process_start)
                    _log(f"Processo abortado em {duration} (minutos:segundos).")
                    return 1
                _log("Validacao final na rede: OK.")
        duration = _format_duration(time.monotonic() - process_start)
        _log(f"Processo finalizado em {duration} (minutos:segundos).")
        return 0
    except KeyboardInterrupt:
        _log("Execucao interrompida pelo usuario (Ctrl+C).")
        duration = _format_duration(time.monotonic() - process_start)
        _log(f"Processo interrompido em {duration} (minutos:segundos).")
        return 130
    except PermissionError as exc:
        detail = redact_secrets(str(exc))
        _log(
            "ERRO ao gravar arquivos. "
            "Possivel causa: planilha/CSV abertos no Excel ou permissao insuficiente."
        )
        if detail:
            _log(f"Detalhe: {exc.__class__.__name__} {detail}")
        else:
            _log(f"Detalhe: {exc.__class__.__name__}")
        duration = _format_duration(time.monotonic() - process_start)
        _log(f"Processo abortado em {duration} (minutos:segundos).")
        return 1
    except Exception as exc:
        detail = _error_detail("Erro inesperado", exc)
        _log(f"ERRO: {detail}")
        duration = _format_duration(time.monotonic() - process_start)
        _log(f"Processo abortado em {duration} (minutos:segundos).")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
