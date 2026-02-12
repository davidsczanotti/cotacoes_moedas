from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
import csv

from openpyxl import Workbook, load_workbook

from cotacoes_moedas.bcb_ptax import PtaxQuote
from cotacoes_moedas.investing import Quote
from cotacoes_moedas.valor_globo import BidAskQuote
from cotacoes_moedas.storage import (
    update_csv_from_xlsx,
    update_xlsx_log,
    update_xlsx_quotes_and_log,
    update_xlsx_usd_brl,
)


def _close_workbook(workbook) -> None:
    close = getattr(workbook, "close", None)
    if callable(close):
        close()


def _make_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Data"
    sheet["B1"] = "Dolar Oficial Compra"
    sheet["C1"] = "Dolar Oficial Venda"
    sheet["L1"] = "TJLP"
    sheet["M1"] = "SELIC"
    sheet["N1"] = "CDI"
    sheet["O1"] = "Situacao"
    workbook.save(path)
    _close_workbook(workbook)


def test_update_xlsx_log_error_format(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "cotacoes.xlsx"
    _make_workbook(xlsx_path)

    target_date = date(2026, 1, 23)
    logged_at = datetime(2026, 1, 23, 9, 15, 0)
    update_xlsx_log(
        xlsx_path,
        target_date=target_date,
        logged_at=logged_at,
        status="ERRO",
        detail="ptax_usd: Timeout",
    )

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    assert sheet["O3"].value == "ERRO 23/01/2026 09:15:00 - ptax_usd: Timeout"
    _close_workbook(workbook)


def test_update_csv_from_xlsx_replaces_same_date(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "cotacoes.xlsx"
    csv_path = tmp_path / "cotacoes.csv"
    _make_workbook(xlsx_path)

    local_tz = datetime.now().astimezone().tzinfo or timezone.utc
    collected_at = datetime(2026, 1, 23, 12, 0, 0, tzinfo=local_tz)
    quote = Quote(
        symbol="USD/BRL",
        value=Decimal("5.2849"),
        value_raw="5,2849",
        collected_at=collected_at,
    )
    update_xlsx_usd_brl(xlsx_path, quote, spread=Decimal("0.0020"))
    update_xlsx_log(xlsx_path, target_date=collected_at.date())
    update_csv_from_xlsx(xlsx_path, csv_path)

    updated_quote = Quote(
        symbol="USD/BRL",
        value=Decimal("5.3001"),
        value_raw="5,3001",
        collected_at=collected_at,
    )
    update_xlsx_usd_brl(xlsx_path, updated_quote, spread=Decimal("0.0020"))
    update_xlsx_log(xlsx_path, target_date=collected_at.date())
    update_csv_from_xlsx(xlsx_path, csv_path)

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))

    data_rows = [row for row in rows if row and row[0] == "23/01/2026"]
    assert len(data_rows) == 1
    assert data_rows[0][1] == "5,3001"
    assert data_rows[0][2] == "5,3021"
    assert rows[0][11:] == ["TJLP", "SELIC", "CDI", "Situacao"]


def test_update_csv_from_xlsx_reads_legacy_log_column(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "cotacoes.xlsx"
    csv_path = tmp_path / "cotacoes.csv"

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Data"
    sheet["L1"] = "Log"
    sheet["A3"] = date(2026, 1, 23)
    sheet["L3"] = "OK 23/01/2026 16:00:00"
    workbook.save(xlsx_path)
    _close_workbook(workbook)

    update_csv_from_xlsx(xlsx_path, csv_path)

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))

    assert rows[1][0] == "23/01/2026"
    assert rows[1][14] == "OK 23/01/2026 16:00:00"


def test_update_xlsx_quotes_and_log_fills_only_blanks(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "cotacoes.xlsx"
    _make_workbook(xlsx_path)

    target_date = date(2026, 1, 23)

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    sheet["A3"] = target_date
    sheet["A3"].number_format = "dd/mm/yyyy"
    sheet["B3"] = Decimal("5.0000")
    sheet["B3"].number_format = "0.0000"
    workbook.save(xlsx_path)
    _close_workbook(workbook)

    collected_at = datetime(2026, 1, 23, 12, 0, 0, tzinfo=timezone.utc)
    usd = Quote(
        symbol="USD/BRL",
        value=Decimal("5.2849"),
        value_raw="5,2849",
        collected_at=collected_at,
    )
    ptax_usd = PtaxQuote(
        symbol="USD/BRL PTAX",
        buy=Decimal("5.1000"),
        sell=Decimal("5.2000"),
        buy_raw="5,1000",
        sell_raw="5,2000",
        collected_at=collected_at,
    )
    turismo = BidAskQuote(
        symbol="USD/BRL Turismo",
        buy=Decimal("5.5000"),
        sell=Decimal("5.6000"),
        buy_raw="5,5000",
        sell_raw="5,6000",
        collected_at=collected_at,
    )

    written = update_xlsx_quotes_and_log(
        xlsx_path,
        target_date=target_date,
        usd_brl=usd,
        ptax_usd=ptax_usd,
        turismo=turismo,
        tjlp=Decimal("9.19"),
        selic=Decimal("15.00"),
        cdi=Decimal("0.0551310642"),
        spread=Decimal("0.0020"),
        overwrite_quotes=False,
        logged_at=datetime(2026, 1, 23, 9, 15, 0),
    )

    assert written["usd_brl"] == ("venda",)
    assert written["ptax_usd"] == ("compra", "venda")
    assert written["turismo"] == ("compra", "venda")
    assert written["tjlp"] == ("valor",)
    assert written["selic"] == ("selic", "cdi")

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    assert Decimal(str(sheet["B3"].value)).quantize(Decimal("0.0001")) == Decimal("5.0000")
    assert Decimal(str(sheet["C3"].value)).quantize(Decimal("0.0001")) == Decimal("5.0020")
    assert Decimal(str(sheet["D3"].value)).quantize(Decimal("0.0001")) == Decimal("5.1000")
    assert Decimal(str(sheet["E3"].value)).quantize(Decimal("0.0001")) == Decimal("5.2000")
    assert Decimal(str(sheet["F3"].value)).quantize(Decimal("0.0001")) == Decimal("5.5000")
    assert Decimal(str(sheet["G3"].value)).quantize(Decimal("0.0001")) == Decimal("5.6000")
    assert Decimal(str(sheet["L3"].value)).quantize(Decimal("0.0001")) == Decimal("0.0919")
    assert Decimal(str(sheet["M3"].value)).quantize(Decimal("0.0001")) == Decimal("0.1500")
    assert Decimal(str(sheet["N3"].value)).quantize(Decimal("0.0000000001")) == Decimal("0.0551310642")
    assert sheet["O3"].value == "OK 23/01/2026 09:15:00"
    assert sheet["L1"].value is None
    assert sheet["M1"].value is None
    assert sheet["N1"].value is None
    assert sheet["O1"].value is None
    assert sheet["L2"].value == "TJLP"
    assert sheet["M2"].value == "SELIC"
    assert sheet["N2"].value == "CDI"
    assert sheet["O2"].value == "Situação"
    merged = {str(ref) for ref in sheet.merged_cells.ranges}
    assert "B1:C1" in merged
    assert "D1:E1" in merged
    assert "F1:G1" in merged
    assert "H1:I1" in merged
    assert "J1:K1" in merged
    assert sheet.auto_filter.ref == "A2:O3"
    assert sheet.freeze_panes == "A3"
    _close_workbook(workbook)


def test_update_xlsx_quotes_and_log_overwrites_when_enabled(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "cotacoes.xlsx"
    _make_workbook(xlsx_path)

    target_date = date(2026, 1, 23)

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    sheet["A3"] = target_date
    sheet["A3"].number_format = "dd/mm/yyyy"
    sheet["B3"] = Decimal("4.0000")
    sheet["B3"].number_format = "0.0000"
    sheet["C3"] = Decimal("4.0020")
    sheet["C3"].number_format = "0.0000"
    workbook.save(xlsx_path)
    _close_workbook(workbook)

    collected_at = datetime(2026, 1, 23, 12, 0, 0, tzinfo=timezone.utc)
    usd = Quote(
        symbol="USD/BRL",
        value=Decimal("5.2849"),
        value_raw="5,2849",
        collected_at=collected_at,
    )

    written = update_xlsx_quotes_and_log(
        xlsx_path,
        target_date=target_date,
        usd_brl=usd,
        spread=Decimal("0.0020"),
        overwrite_quotes=True,
        logged_at=datetime(2026, 1, 23, 9, 15, 0),
    )

    assert written["usd_brl"] == ("compra", "venda")

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    assert Decimal(str(sheet["B3"].value)).quantize(Decimal("0.0001")) == Decimal("5.2849")
    assert Decimal(str(sheet["C3"].value)).quantize(Decimal("0.0001")) == Decimal("5.2869")
    _close_workbook(workbook)


def test_update_xlsx_quotes_and_log_repeats_last_interest_values(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "cotacoes.xlsx"
    _make_workbook(xlsx_path)

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    sheet["A3"] = date(2026, 1, 22)
    sheet["A3"].number_format = "dd/mm/yyyy"
    sheet["L3"] = Decimal("0.0919")
    sheet["M3"] = Decimal("0.1500")
    sheet["N3"] = Decimal("0.0551310642")
    workbook.save(xlsx_path)
    _close_workbook(workbook)

    target_date = date(2026, 1, 23)
    collected_at = datetime(2026, 1, 23, 12, 0, 0, tzinfo=timezone.utc)
    usd = Quote(
        symbol="USD/BRL",
        value=Decimal("5.2849"),
        value_raw="5,2849",
        collected_at=collected_at,
    )

    written = update_xlsx_quotes_and_log(
        xlsx_path,
        target_date=target_date,
        usd_brl=usd,
        spread=Decimal("0.0020"),
        overwrite_quotes=False,
        logged_at=datetime(2026, 1, 23, 9, 15, 0),
    )

    assert written["tjlp"] == ("valor_repetido",)
    assert written["selic"] == ("selic_repetido", "cdi_repetido")

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    assert Decimal(str(sheet["L4"].value)).quantize(Decimal("0.0001")) == Decimal("0.0919")
    assert Decimal(str(sheet["M4"].value)).quantize(Decimal("0.0001")) == Decimal("0.1500")
    assert Decimal(str(sheet["N4"].value)).quantize(Decimal("0.0000000001")) == Decimal("0.0551310642")
    _close_workbook(workbook)
